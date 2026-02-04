#!/usr/bin/env python
# Version: 0.4.1
# Created: 2024-04-07
# Author: ["Hanyuan Zhang"]

import io
from typing import Any, Dict, List, Optional, Union

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pandas.api.types import is_numeric_dtype
from termcolor import cprint

from .calibrator import IsotonicCalibrator, PlattCalibrator
from .metrics import Metrics


class Reporter:
    """
    Comprehensive report generator for machine learning model analysis and validation.

    The Reporter class creates detailed Excel reports containing various aspects of model
    performance, feature analysis, calibration metrics, and data quality assessments.
    It provides formatted Excel output with proper styling for professional reporting.

    Key Features:
    - **Sample Overview**: Statistical summaries and performance metrics across samples
    - **Feature Analysis**: Detailed feature statistics, WOE binning, and quality metrics
    - **Feature Selection**: Documentation of feature selection pipeline and criteria
    - **Model Performance**: Hyperparameter tuning results and feature importance
    - **Calibration Analysis**: Score distribution, PSI tracking, and scorecard generation
    - **Professional Formatting**: Automatic Excel formatting with proper column widths and number formats

    Attributes:
        report_path: Directory path where generated reports will be saved.
                    If None, reports will be saved in the current working directory.

    Example:
        >>> # Initialize reporter with custom path
        >>> reporter = Reporter(report_path='/path/to/reports')
        >>>
        >>> # Generate comprehensive model report
        >>> performance_data = {
        ...     'model_id': 'risk_model_v1',
        ...     'predictions': {'train': train_pred_df, 'test': test_pred_df},
        ...     'label': 'target',
        ...     'woe_df': {
        ...         'train': {'binning': train_binning_df, 'summary': train_summary_df},
        ...         'test': {'binning': test_binning_df, 'summary': test_summary_df}
        ...     },
        ...     'feature_selection': selection_pipeline,
        ...     'calibrate_detail': calibration_metrics
        ... }
        >>> reporter.generate_report(performance_data)
    """

    def __init__(self, report_file_name: Optional[str] = None) -> None:
        """
        Initialize the Reporter with optional custom report directory.

        Args:
            report_file_name: File name for the generated report. If None, a default name will be used.
        """
        self.report_file_name = report_file_name or "model_report.xlsx"

    @classmethod
    def _set_col_format(
        cls, worksheet: Worksheet, col_ids: List[str], format: str
    ) -> None:
        for col_idx in col_ids:
            for cell in worksheet[col_idx]:
                cell.number_format = format

    @classmethod
    def _set_col_width(
        cls, worksheet: Worksheet, col_ids: List[str], width: Union[int, float]
    ) -> None:
        for col_idx in col_ids:
            worksheet.column_dimensions[col_idx].width = width

    def _format_overview_stats_df(self, worksheet: Worksheet) -> None:
        perc_cols = ["C", "D", "E"]
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w20 = ["A", "B", "C", "D", "E"]
        Reporter._set_col_width(worksheet, w20, 20)

    def _format_overview_perf_df(self, worksheet: Worksheet) -> None:
        perc_cols = ["B", "C", "D", "E", "F", "G"]
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w20 = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
        Reporter._set_col_width(worksheet, w20, 20)

    def _format_feat_df(self, worksheet: Worksheet) -> None:
        perc_cols = ["D", "F", "G", "H", "S", "U", "W", "X", "Y"]
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w40 = ["A"]
        w25 = ["G", "H", "I", "J"]
        w20 = [
            "B",
            "C",
            "D",
            "E",
            "F",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
        ]
        w15 = ["X", "Y"]
        Reporter._set_col_width(worksheet, w40, 40)
        Reporter._set_col_width(worksheet, w25, 25)
        Reporter._set_col_width(worksheet, w20, 20)
        Reporter._set_col_width(worksheet, w15, 10)

    def _format_woe_df(self, worksheet: Worksheet) -> None:
        num_cols_4dec = ["N", "O", "Q"]
        perc_cols = ["D", "E", "G", "H", "I", "K", "L", "M", "P", "S", "T"]
        Reporter._set_col_format(worksheet, num_cols_4dec, "0.0000")
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w40 = ["A"]
        w20 = ["B"]
        w15 = [get_column_letter(i) for i in range(3, worksheet.max_column + 1)]
        Reporter._set_col_width(worksheet, w40, 40)
        Reporter._set_col_width(worksheet, w20, 20)
        Reporter._set_col_width(worksheet, w15, 15)

    def _format_feature_selection_overview_df(self, worksheet: Worksheet) -> None:
        w40 = ["A"]
        Reporter._set_col_width(worksheet, w40, 40)

    def _format_feature_selection_df(self, worksheet: Worksheet) -> None:
        w20 = ["C", "D", "E"]
        w40 = ["B"]
        Reporter._set_col_width(worksheet, w20, 20)
        Reporter._set_col_width(worksheet, w40, 40)

    def _format_tuning_df(self, worksheet: Worksheet) -> None:
        perc_cols = ["P", "Q", "R", "S", "T", "U", "V", "W", "X"]
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w15 = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "T",
            "U",
            "V",
            "W",
            "X",
        ]
        w20 = ["P", "Q", "R", "S"]
        w40 = ["L", "M", "N", "O"]
        Reporter._set_col_width(worksheet, w15, 15)
        Reporter._set_col_width(worksheet, w20, 20)
        Reporter._set_col_width(worksheet, w40, 40)

    def _format_calibration_reg_df(self, worksheet: Worksheet) -> None:
        w15 = ["A", "B", "C", "D"]
        w20 = ["E", "F"]
        Reporter._set_col_width(worksheet, w15, 15)
        Reporter._set_col_width(worksheet, w20, 20)

    def _format_calibration_scorecard_df(self, worksheet: Worksheet) -> None:
        perc_cols = [
            "D",
            "E",
            "G",
            "H",
            "I",
            "J",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
        ]
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w15 = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
            "U",
        ]
        Reporter._set_col_width(worksheet, w15, 15)

    def _format_calibration_scoredist_df(self, worksheet: Worksheet) -> None:
        perc_cols = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U"]
        Reporter._set_col_format(worksheet, perc_cols, "0.000%")
        w15 = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U"]
        Reporter._set_col_width(worksheet, w15, 15)

    def _format_calibration_score_psi_df(self, worksheet: Worksheet) -> None:
        perc_cols = ["B"]
        Reporter._set_col_format(worksheet, perc_cols, "0.00%")
        w15 = ["B"]
        Reporter._set_col_width(worksheet, w15, 15)
        w20 = ["A"]
        Reporter._set_col_width(worksheet, w20, 20)

    def _insert_figure_to_excel(
        self,
        writer: pd.ExcelWriter,
        fig: plt.Figure,
        sheet_name: str,
        anchor: str = "A3",
        title: str = "",
        dpi: int = 50,
    ) -> None:
        # Insert matplotlib figure into Excel worksheet as PNG image
        # Create a placeholder DataFrame for the sheet
        header_text = title if title else sheet_name
        pd.DataFrame({sheet_name: [header_text]}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )

        worksheet = writer.sheets[sheet_name]

        # Save figure to byte buffer and add to worksheet
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
        buf.seek(0)
        img = XLImage(buf)
        img.anchor = anchor
        worksheet.add_image(img)

    def _stat_perf(self, gp: pd.DataFrame, target_label: str) -> pd.DataFrame:
        # Calculate AUC, KS, Gini, IV metrics for main model predictions
        y_true = gp[target_label]
        y_proba = gp["proba"]
        if y_true.nunique() == 1:
            return pd.DataFrame(
                {
                    "AUC": [None],
                    "KS": [None],
                    "Gini": [None],
                    "IV": [None],
                }
            )
        return pd.DataFrame(
            {
                "AUC": [Metrics.get_auc(y_true, y_proba)],
                "KS": [Metrics.get_ks(y_true, y_proba)],
                "Gini": [Metrics.get_gini(y_true, y_proba)],
                "IV": [Metrics.get_iv(y_true, y_proba)],
            }
        )

    def _generate_sample_overview_basic(
        self, writer: pd.ExcelWriter, predictions: Dict[str, pd.DataFrame]
    ) -> None:
        # Generate basic sample statistics (size, bad rate) without model performance metrics
        sample_stats = []
        for sample_type, pred_df in predictions.items():
            y = pred_df["target"]
            sample_stats.append(
                {
                    "sample_type": sample_type,
                    "# Sample": len(y),
                    "% Bad": y.mean() if hasattr(y, "mean") else sum(y) / len(y),
                }
            )
        df_basic = pd.DataFrame(sample_stats).set_index("sample_type")
        df_basic["% Sample"] = df_basic["# Sample"] / df_basic["# Sample"].sum()
        df_basic = df_basic[["# Sample", "% Sample", "% Bad"]]
        df_basic.sort_index(ascending=False).to_excel(
            writer, sheet_name="Sample Overview", freeze_panes=(1, 1)
        )

    def generate_sample_overview_report(
        self, writer: pd.ExcelWriter, predictions: Dict[str, pd.DataFrame], label: str
    ) -> None:
        """
        Generate comprehensive sample overview report with statistics and performance metrics.

        Creates two worksheets containing sample-level analysis:
        1. **Sample Overview - Statistics**: Sample sizes, bad rates, and PSI values
        2. **Sample Overview - Performance**: Model performance metrics by sample

        Args:
            writer: Excel writer object for output file generation.
            predictions: Dictionary containing prediction DataFrames by sample type.
                Each DataFrame must contain 'sample_type' and 'score' columns.
            label: Target variable column name for bad rate calculations.

        Raises:
            KeyError: If required columns ('sample_type', 'score') are missing from predictions.
            ValueError: If predictions structure is invalid or empty.

        Example:
            >>> # Generate sample overview
            >>> predictions = {'train': train_pred_df, 'test': test_pred_df}
            >>> with pd.ExcelWriter('report.xlsx') as writer:
            ...     reporter.generate_sample_overview_report(writer, predictions, 'target')
        """
        df_extra = pd.concat([pred_df for _, pred_df in predictions.items()])
        df_basic_summary = df_extra.groupby("sample_type").agg(
            sample_size=("sample_type", "size"),
            sample_size_pct=("sample_type", lambda x: x.size / df_extra.shape[0]),
            positive_rate=(label, "mean"),
        )
        df_basic_summary.columns = ["# Sample", "% Sample", "% Bad"]
        df_basic_summary.sort_index(ascending=False).to_excel(
            writer, sheet_name="Sample Overview - Statistics", freeze_panes=(1, 1)
        )
        self._format_overview_stats_df(writer.sheets["Sample Overview - Statistics"])

        df_extra.groupby("sample_type").apply(
            self._stat_perf, target_label=label
        ).droplevel(level=1).sort_index(ascending=False).to_excel(
            writer, sheet_name="Sample Overview - Performance", freeze_panes=(1, 1)
        )
        self._format_overview_perf_df(writer.sheets["Sample Overview - Performance"])

    def generate_single_feature_eda_report(
        self,
        writer: pd.ExcelWriter,
        woe_data: Dict[str, pd.DataFrame],
        prefix: str = "",
    ) -> None:
        """
        Generate comprehensive Exploratory Data Analysis (EDA) report for features.

        Creates detailed feature analysis worksheets containing:
        1. **Feature Overview**: Comprehensive feature statistics and quality metrics
        2. **Feature Binning Report**: Weight of Evidence binning analysis with KS/IV

        Args:
            writer: Excel writer object for output file generation.
            woe_data: Dictionary containing:
                - 'binning': Bin-level WOE DataFrame with KS and IV metrics
                - 'summary': Feature-level statistics DataFrame
            prefix: Optional prefix for worksheet names (e.g., "Train", "Test").

        Raises:
            KeyError: If woe_data doesn't contain required keys ('binning', 'summary').
            ValueError: If DataFrames don't contain required columns or structure.

        Example:
            >>> # Generate feature EDA report
            >>> woe_data = encoder.get_woe_df(X_train, y_train)
            >>> with pd.ExcelWriter('feature_analysis.xlsx') as writer:
            ...     reporter.generate_single_feature_eda_report(
            ...         writer, woe_data, "Train"
            ...     )
        """
        binning_df = woe_data["binning"]
        summary_df = woe_data["summary"]

        summary_df.to_excel(
            writer, sheet_name=f"{prefix} - Feature Overview", freeze_panes=(1, 1)
        )
        binning_df.to_excel(
            writer, sheet_name=f"{prefix} - Feature Binning Report", freeze_panes=(3, 2)
        )
        self._format_feat_df(writer.sheets[f"{prefix} - Feature Overview"])
        self._format_woe_df(writer.sheets[f"{prefix} - Feature Binning Report"])

    def generate_feature_selection_report(
        self, writer: pd.ExcelWriter, fitted_steps: Any
    ) -> None:
        """
        Generate comprehensive feature selection pipeline report.

        Creates detailed analysis of the feature selection process including:
        1. **Feature Selection - Overview**: Summary showing which features survived each selection step
        2. **Feature Selection - [Method]**: Detailed results for each selection method

        Args:
            writer: Excel writer object for output file generation.
            fitted_steps: Fitted steps from Preprocess. Can be either:
                - Dict[str, transformer]: From preprocessor.named_steps
                - List[Tuple[str, transformer]]: From preprocessor._fitted_steps
                Each selector should have 'selected_features' and 'detail' attributes.

        Selection Methods Supported:
        - **Correlation (Corr)**: Uses selector.detail['after'] for results
        - **Other Methods**: Uses selector.detail directly for results

        Example:
            >>> # Generate feature selection report
            >>> with pd.ExcelWriter('feature_selection.xlsx') as writer:
            ...     reporter.generate_feature_selection_report(writer, preprocessor.named_steps)
        """
        # Normalize input to list of tuples
        if isinstance(fitted_steps, dict):
            steps_list = list(fitted_steps.items())
        else:
            steps_list = list(fitted_steps)

        # Filter to only selector steps (by checking module and attribute)
        selector_steps = [
            (name, step)
            for name, step in steps_list
            if "feature_selection" in getattr(step.__class__, "__module__", "")
            and hasattr(step, "selected_features")
        ]

        if not selector_steps:
            return

        # Find "Original" step for baseline features
        original_features = None
        for name, selector in selector_steps:
            if name == "Original":
                original_features = selector.selected_features
                break

        # Fallback: use first selector's features as baseline
        if original_features is None:
            original_features = selector_steps[0][1].selected_features

        # Handle empty features case
        if original_features is None or len(original_features) == 0:
            # Create an empty overview sheet with a message
            pd.DataFrame(
                {"Message": ["No features available for selection overview"]}
            ).to_excel(writer, sheet_name="Feature Selection - Overview", index=False)
            return

        # Build overview DataFrame
        original_cols = pd.DataFrame(original_features, columns=["feature"]).set_index(
            "feature"
        )

        for name, selector in selector_steps:
            if name == "Original":
                continue
            original_cols[name] = 0
            if (
                selector.selected_features is not None
                and len(selector.selected_features) > 0
            ):
                # Mark selected features as 1
                for feat in selector.selected_features:
                    if feat in original_cols.index:
                        original_cols.loc[feat, name] = 1

        # Sort by selection order (last selector first)
        sort_columns = [name for name, _ in selector_steps if name != "Original"][::-1]
        if sort_columns:
            original_cols.sort_values(sort_columns, ascending=False).to_excel(
                writer, sheet_name="Feature Selection - Overview"
            )
        else:
            original_cols.to_excel(writer, sheet_name="Feature Selection - Overview")

        self._format_feature_selection_overview_df(
            writer.sheets["Feature Selection - Overview"]
        )

        # Write detail sheets for each selector
        for name, selector in selector_steps:
            if name == "Original":
                continue
            if not hasattr(selector, "detail") or selector.detail is None:
                continue

            if name == "Corr" and isinstance(selector.detail, dict):
                detail_df = selector.detail.get("after")
                if detail_df is not None and not detail_df.empty:
                    detail_df.to_excel(
                        writer,
                        sheet_name=f"Feature Selection - {name}",
                        freeze_panes=(1, 1),
                    )
            elif (
                isinstance(selector.detail, pd.DataFrame) and not selector.detail.empty
            ):
                selector.detail.to_excel(
                    writer,
                    sheet_name=f"Feature Selection - {name}",
                    freeze_panes=(1, 1),
                )
                self._format_feature_selection_df(
                    writer.sheets[f"Feature Selection - {name}"]
                )

    def generate_model_tuning_report(
        self, writer: pd.ExcelWriter, tune_results: Optional[pd.DataFrame]
    ) -> None:
        """
        Generate hyperparameter tuning results report.

        Creates detailed analysis of hyperparameter optimization including parameter
        combinations tested, performance metrics, and optimal parameter selection.

        Args:
            writer: Excel writer object for output file generation.
            tune_results: DataFrame containing tuning results with hyperparameters and metrics.
                         If None, no tuning report will be generated.

        Example:
            >>> # Generate tuning report (if tuning was performed)
            >>> with pd.ExcelWriter('tuning_results.xlsx') as writer:
            ...     if tune_results is not None:
            ...         reporter.generate_model_tuning_report(writer, tune_results)
        """
        if tune_results is not None:
            tune_results.to_excel(
                writer, sheet_name="Model - Tuning Report", freeze_panes=(1, 1)
            )
            self._format_tuning_df(writer.sheets["Model - Tuning Report"])

    def generate_feature_importance_report(
        self,
        writer: pd.ExcelWriter,
        feature_importance: pd.DataFrame,
    ) -> None:
        """
        Generate feature importance report with visualization.

        Creates a detailed report of feature importance from the trained model,
        including a bar chart visualization and sorted importance values.

        Args:
            writer: Excel writer object for output file generation.
            feature_importance: DataFrame with 'feature' and 'importance' columns.

        Example:
            >>> # Generate feature importance report
            >>> with pd.ExcelWriter('report.xlsx') as writer:
            ...     reporter.generate_feature_importance_report(writer, feat_imp_df)
        """
        if feature_importance is None or len(feature_importance) == 0:
            return

        # Sort by importance descending
        df_imp = feature_importance.sort_values(
            "importance", ascending=False
        ).reset_index(drop=True)
        df_imp.to_excel(
            writer,
            sheet_name="Model - Feature Importance",
            freeze_panes=(1, 1),
            index=False,
        )

        # Format the worksheet
        worksheet = writer.sheets["Model - Feature Importance"]
        self._set_col_width(worksheet, ["A"], 40)
        self._set_col_width(worksheet, ["B"], 15)
        self._set_col_format(worksheet, ["B"], "0.0000")

        # Create bar chart
        fig, ax = plt.subplots(figsize=(10, max(6, len(df_imp) * 0.3)))
        top_n = min(30, len(df_imp))  # Show top 30 features in chart
        df_plot = df_imp.head(top_n).iloc[::-1]  # Reverse for horizontal bar
        ax.barh(df_plot["feature"], df_plot["importance"], color="steelblue")
        ax.set_xlabel("Importance")
        ax.set_ylabel("Feature")
        ax.set_title(f"Top {top_n} Feature Importance")
        plt.tight_layout()

        # Insert chart into worksheet
        img_buffer = io.BytesIO()
        fig.savefig(img_buffer, format="png", dpi=100, bbox_inches="tight")
        plt.close(fig)
        img_buffer.seek(0)
        img = XLImage(img_buffer)
        img.anchor = "D2"
        worksheet.add_image(img)

    def generate_calibration_report(
        self,
        writer: pd.ExcelWriter,
        calibrator: Union[PlattCalibrator, IsotonicCalibrator],
        scorecard: Dict[str, pd.DataFrame],
        scoredist: Dict[str, pd.DataFrame],
    ) -> None:
        """
        Generate comprehensive model calibration analysis report.

        Creates detailed calibration analysis including regression parameters,
        scorecards for different sample types, score distributions, PSI tracking,
        and optional calibration plot.

        Report Components:
        1. **Calibration - Regression**: Calibration regression parameters and fit metrics
        2. **Calibration - [Sample] Score Card**: Score bands with population and risk metrics
        3. **Calibration - Score Distribution**: Score distribution across samples
        4. **Calibration - Score PSI**: Population Stability Index for score drift monitoring
        5. **Calibration - Plot**: Before/after calibration curves (if available)

        Args:
            writer: Excel writer object for output file generation.
            calibrator: Fitted Calibration instance containing calibrate_detail and
                       optionally calibrate_plot.
            scorecard: Dictionary mapping sample names to scorecard DataFrames:
                      {'train': scorecard_df, 'test': scorecard_df, ...}
            scoredist: Dictionary containing distribution analysis:
                      {'Distribution': dist_df, 'PSI': psi_df}

        Example:
            >>> # Generate calibration report with calibrator (probability mode)
            >>> from optimus.calibrator import Calibration
            >>> calibrator = Calibration()  # probability mode (no mapping_base)
            >>> calibrator.fit(y_prob_train, y_train)
            >>>
            >>> scorecard_data = {'train': train_scorecard, 'test': test_scorecard}
            >>> dist_data = {'Distribution': score_dist, 'PSI': psi_analysis}
            >>> with pd.ExcelWriter('calibration_report.xlsx') as writer:
            ...     reporter.generate_calibration_report(
            ...         writer, calibrator, scorecard_data, dist_data
            ...     )
        """
        if calibrator is None:
            return

        if calibrator.calibrate_detail is not None:
            calibrator.calibrate_detail.to_excel(
                writer, sheet_name="Calibration - Regression", freeze_panes=(1, 1)
            )
            self._format_calibration_reg_df(writer.sheets["Calibration - Regression"])

        for name, df_scorecard in scorecard.items():
            df_scorecard.to_excel(
                writer,
                sheet_name=f"Calibration - {name} Score Card",
                freeze_panes=(1, 1),
            )
            self._format_calibration_scorecard_df(
                writer.sheets[f"Calibration - {name} Score Card"]
            )

        scoredist["Distribution"].to_excel(
            writer,
            sheet_name=f"Calibration - Score Distribution",
            freeze_panes=(3, 1),
        )
        scoredist["PSI"].to_excel(
            writer,
            sheet_name=f"Calibration - Score PSI",
            freeze_panes=(1, 1),
        )
        self._format_calibration_scoredist_df(
            writer.sheets[f"Calibration - Score Distribution"]
        )
        self._format_calibration_score_psi_df(writer.sheets[f"Calibration - Score PSI"])

        # Insert calibration plot if calibrator has a plot
        if calibrator.calibrate_plot is not None:
            # Insert calibration plot with smaller dpi for more compact image
            self._insert_figure_to_excel(
                writer=writer,
                fig=calibrator.calibrate_plot,
                sheet_name="Calibration - Plot",
                title="Calibration Curve (Before vs After)",
                dpi=100,
            )

        if isinstance(calibrator, IsotonicCalibrator):
            try:
                scaling_plot = calibrator.plot_score_vs_calibrated_prob()
                self._insert_figure_to_excel(
                    writer=writer,
                    fig=scaling_plot,
                    sheet_name="Calibration - Scaling Mapping",
                    title="Linear Scaling: Calibrated Probability → Scaled Score",
                    dpi=100,
                )
            except Exception as e:
                cprint(f"[WARN] Failed to generate scaling mapping plot: {e}", "yellow")

    def generate_report(self, performance: Dict[str, Any], **kwargs) -> None:
        """
        Generate comprehensive model performance report with all analysis components.

        Creates a complete Excel report containing all available analysis sections
        based on the provided performance data. Automatically detects whether to generate
        EDA-only or full model report based on the presence of 'label' key.

        Report Sections Generated (based on available data):
        - **Sample Overview**: Always generated
            - Full version (with PSI/Performance) if 'label' is provided
            - Basic version (sizes/bad rates only) if 'label' is not provided
        - **Feature Analysis**: Generated if 'woe_df' provided
        - **Feature Selection**: Generated if 'feature_selection' provided
        - **Model Tuning**: Generated if 'tune_results' provided
        - **Calibration Analysis**: Generated if calibration data provided

        Args:
            performance: Dictionary containing model performance data:
                Required:
                - **predictions** (Dict): Prediction results by sample type
                    Each sample is a DataFrame with predictions and scores

                Required for full report:
                - **label** (str): Target column name in predictions DataFrame
                    If not provided, generates EDA-only report

                Optional:
                - **woe_df** (Dict): WOE analysis by sample type, each containing
                    {'binning': bin_df, 'summary': summary_df}
                - **missing_values** (List): Values to treat as missing (default: [])
                - **feature_selection**: Feature selection pipeline results
                - **tune_results** (DataFrame): Hyperparameter tuning results
                - **calibrate_detail** (DataFrame): Calibration regression details
                - **scorecard** (Dict): Scorecard analysis by sample
                - **scoredist** (Dict): Score distribution and PSI analysis
            **kwargs: Additional keyword arguments (for future extensibility).

        Output:
            Creates Excel file at self.report_file_name

        Example:
            >>> # Full report with model scores
            >>> reporter.generate_report({
            ...     'predictions': {'train': train_pred_df, 'test': test_pred_df},
            ...     'label': 'default_flag',
            ...     'woe_df': {'train': {'binning': train_bin, 'summary': train_sum},
            ...                'test': {'binning': test_bin, 'summary': test_sum}},
            ...     'missing_values': [-999999],
            ...     'feature_selection': preprocessor.named_steps,
            ...     'calibrate_detail': calib_df,
            ...     'scorecard': scorecard_dict,
            ...     'scoredist': scoredist_dict,
            ... })
        """
        writer = pd.ExcelWriter(
            self.report_file_name,
            engine="openpyxl",
        )

        predictions = performance["predictions"]

        # Sample Overview
        if "label" in performance:
            self.generate_sample_overview_report(
                writer, predictions, performance["label"]
            )
        else:
            self._generate_sample_overview_basic(writer, predictions)

        # Feature EDA
        if "woe_df" in performance:
            for sample_type, woe_data in performance["woe_df"].items():
                if sample_type in predictions:
                    self.generate_single_feature_eda_report(
                        writer,
                        woe_data,
                        sample_type,
                    )

        # Feature Selection
        if "feature_selection" in performance:
            self.generate_feature_selection_report(
                writer, performance["feature_selection"]
            )

        # Model Tuning
        if "tune_results" in performance:
            self.generate_model_tuning_report(writer, performance["tune_results"])

        # Feature Importance
        if "feature_importance" in performance:
            self.generate_feature_importance_report(
                writer, performance["feature_importance"]
            )

        if (
            "calibrator" in performance
            and "scorecard" in performance
            and "scoredist" in performance
        ):
            self.generate_calibration_report(
                writer,
                performance["calibrator"],
                performance["scorecard"],
                performance["scoredist"],
            )

        writer.close()
        cprint(f"[INFO] Report saved to: {self.report_file_name}", "green")
